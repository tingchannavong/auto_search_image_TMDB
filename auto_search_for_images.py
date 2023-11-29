import openpyxl
from search_for_images_function import download_movie_image
from search_for_tvshows_function import download_tv_show_image

#CHOOSE DESIRED IMAGE FOLDEER PATH
custom_folder_path = "C:\\Users\\thipphaphone.c\\Desktop\\Learning\\movies_images"

def start_image_search_from(sheetname, start_row, watch_type):

    # Open the Excel file
    excel_file_path = "C:\\Users\\thipphaphone.c\\Desktop\\Learning\\fav_watched_list.xlsx"
    workbook = openpyxl.load_workbook(excel_file_path)
    worksheet = workbook[sheetname]

    #Iterate over list, download image, save image path to desired column
    fixed_column = 1
    current_row = start_row

    for row in worksheet.iter_rows(min_row=current_row, max_col=fixed_column, values_only=True):
        name = row[0]  # Where 0 represents the start of iteration

        # Check if watch_type is valid
        if watch_type == 'series':
            poster_path = download_tv_show_image(name, save_folder=custom_folder_path)
            worksheet.cell(row=current_row, column=4, value=poster_path)
            print(f"Excel file updated with {name} path in column C, row {current_row}")

        elif watch_type == 'movies':
            poster_path = download_movie_image(name,save_folder=custom_folder_path)
            worksheet.cell(row=current_row, column=3, value=poster_path)
            print(f"Excel file updated with {name} path in column C, row {current_row}") 

        else:
            print("Please enter watch_type as: 'movies' or 'series' ONLY")

        #Increment current row
        current_row += 1

    # Save the modified Excel file
    workbook.save(excel_file_path)

start_image_search_from('movies_sheet', 2, 'movies')
